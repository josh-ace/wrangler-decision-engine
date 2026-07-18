"""Analysis_TrimPath — three-price feature transparency + landed cost per trim path.

Two Excel Tables on one tab, per the IC #4 design (``analysis-designs/trim_path.md``):

- ``TrimPathDetail`` — the computational engine, one row per (config x covered
  feature) = 28 x 20 = 560 rows. Each row decides whether the feature is
  standard-embedded on that trim, gates it by the user's Target_Build selection, and
  computes the aftermarket parts/labor it contributes to the landed cost.
- ``TrimPathSummary`` — 28 rows, one per config, that ``SUMIFS`` the detail's
  per-feature contributions into a landed cost (MSRP + aftermarket-to-target) and a
  delta versus the cheapest path.

Every reference is a **structured reference** (``Trims[MSRP]``,
``TrimPathDetail[Config_Key]``, ``Target_Build[Include]``) or a named range
(``Input_Labor_Rate``) — no ``$A$2`` cell addresses, so refresh and row add/remove
never break a formula. Python only writes the strings; Excel computes.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from engine.data import mod_pricing, trims
from engine.xlsx import write_plain_table

TAB_NAME = "Analysis_TrimPath"
TITLE = "Trim path — three-price transparency + landed cost"

DETAIL_TABLE = "TrimPathDetail"
SUMMARY_TABLE = "TrimPathSummary"

# Literal shown wherever a factory-option path would go. Per-trim option pricing
# lives in Data_Options, which is not populated in MVP, so factory-option paths are
# explicitly unavailable rather than stubbed with a fake number (design rec (a)).
FACTORY_OPTION_NA = "N/A (pending Data_Options)"

_DOLLAR_FORMAT = '"$"#,##0'

# Column order for each table (verbatim from the design's "Analysis_TrimPath layout").
DETAIL_COLUMNS = [
    "Trim",
    "Body",
    "Powertrain",
    "Feature",
    "Config_Key",
    "Included",
    "Standard_On_Trim",
    "Factory_Option_Price",
    "Aftermarket_Parts",
    "Aftermarket_Labor",
    "Aftermarket_Total",
    "Parts_Applied",
    "Labor_Applied",
    "Contribution_to_Landed",
]

SUMMARY_COLUMNS = [
    "Trim",
    "Body",
    "Powertrain",
    "Config_Key",
    "Base_MSRP",
    "Factory_Options_Cost",
    "Aftermarket_Cost",
    "Aftermarket_Labor_Cost",
    "Landed_Cost",
    "Delta_vs_Cheapest",
]

# Dollar columns to number-format for legibility, keyed by the table they belong to.
_DETAIL_DOLLAR_COLUMNS = [
    "Aftermarket_Parts",
    "Aftermarket_Labor",
    "Aftermarket_Total",
    "Parts_Applied",
    "Labor_Applied",
    "Contribution_to_Landed",
]
_SUMMARY_DOLLAR_COLUMNS = [
    "Base_MSRP",
    "Aftermarket_Cost",
    "Aftermarket_Labor_Cost",
    "Landed_Cost",
    "Delta_vs_Cheapest",
]

# Config_Key expression, identical in both tables (each references its own [@ cols).
_CONFIG_KEY = '=[@Trim]&" / "&[@Body]&" / "&[@Powertrain]'


def _configs() -> list[tuple[str, str, str]]:
    """The 28 (Trim, Body, Powertrain) identity tuples, in Data_Trims order.

    Written as literal enumeration into both tables — they are stable join keys from
    the order guide, not computed values, so a refresh that changes an MSRP flows
    through the formulas while these identity columns stay put.
    """
    return [(row[0], row[1], row[2]) for row in trims.load()]


def _feature_names() -> list[str]:
    """The 20 mod-pricing-covered feature display names, in Data_ModPricing order.

    ``mod_pricing.load()`` already emits the taxonomy display name in its Feature
    column — the same name Features / ModPricing / Target_Build all join by — so
    reuse it rather than re-deriving the id->name map.
    """
    return [row[0] for row in mod_pricing.load()]


def _detail_rows() -> list[list]:
    """Build the 560 TrimPathDetail rows (config-major, one row per feature).

    Identity columns (Trim/Body/Powertrain/Feature) are literals; everything else is
    an Excel formula string using current-row structured references (``[@Col]``).
    """
    rows: list[list] = []
    for trim, body, powertrain in _configs():
        for feature in _feature_names():
            rows.append(
                [
                    trim,
                    body,
                    powertrain,
                    feature,
                    _CONFIG_KEY,
                    # Included — user's Target_Build gate; missing feature -> "No".
                    '=IFERROR(INDEX(Target_Build[Include], '
                    'MATCH([@Feature], Target_Build[Feature], 0)), "No")',
                    # Standard_On_Trim — delimiter-bracketed membership test so
                    # "Sport" does not false-match inside "Sport S". Returns a boolean.
                    "=IFERROR(ISNUMBER(SEARCH("
                    '", "&[@Trim]&",", '
                    '", "&INDEX(Features[Standard_On], '
                    'MATCH([@Feature], Features[Feature], 0))&",")), FALSE)',
                    # Factory_Option_Price — N/A in MVP (no Data_Options).
                    FACTORY_OPTION_NA,
                    # Aftermarket_Parts — join ModPricing by feature name.
                    "=IFERROR(INDEX(ModPricing[Parts_Cost], "
                    'MATCH([@Feature], ModPricing[Feature], 0)), "")',
                    # Aftermarket_Labor — install hours x the labor-rate named range.
                    "=IFERROR(INDEX(ModPricing[Install_Hours], "
                    'MATCH([@Feature], ModPricing[Feature], 0)) * Input_Labor_Rate, "")',
                    # Aftermarket_Total — parts + labor.
                    '=IFERROR([@Aftermarket_Parts] + [@Aftermarket_Labor], "")',
                    # Parts_Applied — 0 if not included, 0 if standard, else parts.
                    '=IF([@Included] <> "Yes", 0, '
                    "IF([@Standard_On_Trim], 0, IFERROR([@Aftermarket_Parts], 0)))",
                    # Labor_Applied — same gate, on labor.
                    '=IF([@Included] <> "Yes", 0, '
                    "IF([@Standard_On_Trim], 0, IFERROR([@Aftermarket_Labor], 0)))",
                    # Contribution_to_Landed — parts + labor actually applied.
                    "=[@Parts_Applied] + [@Labor_Applied]",
                ]
            )
    return rows


def _summary_rows() -> list[list]:
    """Build the 28 TrimPathSummary rows, one per config."""
    rows: list[list] = []
    for trim, body, powertrain in _configs():
        rows.append(
            [
                trim,
                body,
                powertrain,
                _CONFIG_KEY,
                # Base_MSRP — keyed join into Trims (robust to row reordering/refresh).
                "=SUMIFS(Trims[MSRP], Trims[Trim], [@Trim], "
                "Trims[Body], [@Body], Trims[Powertrain], [@Powertrain])",
                # Factory_Options_Cost — N/A in MVP; excluded from Landed_Cost.
                FACTORY_OPTION_NA,
                # Aftermarket_Cost — sum the detail's applied parts for this config.
                "=SUMIFS(TrimPathDetail[Parts_Applied], "
                "TrimPathDetail[Config_Key], [@Config_Key])",
                # Aftermarket_Labor_Cost — sum the detail's applied labor.
                "=SUMIFS(TrimPathDetail[Labor_Applied], "
                "TrimPathDetail[Config_Key], [@Config_Key])",
                # Landed_Cost — MSRP + aftermarket parts + aftermarket labor.
                "=[@Base_MSRP] + [@Aftermarket_Cost] + [@Aftermarket_Labor_Cost]",
                # Delta_vs_Cheapest — vs the cheapest of all 28 landed paths.
                "=[@Landed_Cost] - MIN(TrimPathSummary[Landed_Cost])",
            ]
        )
    return rows


def _format_dollars(
    ws: Worksheet, columns: list[str], dollar_columns: list[str], header_row: int, n_rows: int
) -> None:
    """Apply the dollar number format to the given columns' data cells."""
    for name in dollar_columns:
        col_idx = columns.index(name) + 1
        for r in range(header_row + 1, header_row + 1 + n_rows):
            ws.cell(row=r, column=col_idx).number_format = _DOLLAR_FORMAT


def setup_formulas(ws: Worksheet) -> None:
    """Write the two-table trim-path build onto ``ws``.

    Layout: title (row 1), Section 1 label + TrimPathSummary (header row 4), then
    Section 2 label + TrimPathDetail (header row 37). Both tables enumerate the same
    28 configs Python reads from the order guide at render time.
    """
    ws["A1"] = TITLE

    # --- Section 1: landed cost per trim path (the summary) ---
    ws["A3"] = "Section 1 — Landed cost per trim path"
    summary_header_row = 4
    summary_rows = _summary_rows()
    write_plain_table(
        ws, SUMMARY_TABLE, SUMMARY_COLUMNS, summary_rows, start_row=summary_header_row
    )
    _format_dollars(
        ws, SUMMARY_COLUMNS, _SUMMARY_DOLLAR_COLUMNS, summary_header_row, len(summary_rows)
    )

    # --- Section 2: per-feature breakdown (three-price detail) ---
    ws["A36"] = "Section 2 — Per-feature breakdown (three-price detail)"
    detail_header_row = 37
    detail_rows = _detail_rows()
    write_plain_table(
        ws, DETAIL_TABLE, DETAIL_COLUMNS, detail_rows, start_row=detail_header_row
    )
    _format_dollars(
        ws, DETAIL_COLUMNS, _DETAIL_DOLLAR_COLUMNS, detail_header_row, len(detail_rows)
    )
