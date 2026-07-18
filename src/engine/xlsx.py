"""Shared openpyxl helpers that enforce the Excel-Table / named-range discipline.

Every Python-owned ``Data_*`` tab stores its rows inside a real Excel Table so
that `engine refresh` can rewrite rows without breaking any downstream formula
(structured references like ``Trims[MSRP]`` survive row add/remove; ``$A$2:$A$20``
would not). This module is the single place that knows how to lay a table down
and how to clear one for a refresh, so every data module stays a thin declaration.
"""

from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

# Provenance columns appended to every Data_* table. Per spec's MVP provenance
# framework, every data value carries where it came from and when it was pulled.
PROVENANCE_COLUMNS = ["Source", "As_Of_Date"]

_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)


def write_data_table(
    ws: Worksheet,
    table_name: str,
    columns: list[str],
    rows: list[list] | None = None,
) -> Table:
    """Write ``columns`` as a header row plus ``rows`` (or one empty placeholder
    row) and wrap the whole block in an Excel Table named ``table_name``.

    An Excel Table needs at least one data row for its reference range, so when a
    module has no data yet (the scaffolding case) we emit a single blank row.
    Future ICs return real rows from ``load()`` and this function grows the table
    downward — the table name, and therefore every structured reference, is stable.
    """
    header = [*columns, *PROVENANCE_COLUMNS]

    for col_idx, name in enumerate(header, start=1):
        ws.cell(row=1, column=col_idx, value=name)

    data_rows = rows if rows else [[None] * len(header)]
    for row_offset, row in enumerate(data_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_offset, column=col_idx, value=value)

    last_col = get_column_letter(len(header))
    last_row = 1 + len(data_rows)
    ref = f"A1:{last_col}{last_row}"

    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = _TABLE_STYLE
    ws.add_table(table)

    # Sensible starting widths so the placeholder tabs are legible on open.
    for col_idx, name in enumerate(header, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(name) + 2)

    return table


def clear_data_table(ws: Worksheet, table_name: str) -> None:
    """Drop the named table and wipe the sheet so a data module can re-render it.

    Used by `engine refresh`: the table object and its cells are removed, then the
    module's ``render()`` lays a fresh table down. This is why refresh touches only
    ``Data_*`` tabs — user tabs and Analysis tabs are never handed to this path.
    """
    if table_name in ws.tables:
        del ws.tables[table_name]
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)
