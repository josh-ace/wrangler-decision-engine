"""Workbook assembly — the render orchestration.

``build_workbook()`` lays down the full tab structure from the module registries
(``DATA_MODULES`` / ``ANALYSIS_MODULES``) plus the user tabs, reference tabs, and
named ranges. This is the one place that knows the overall shape of the report;
individual tabs know only themselves.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from engine.analysis import ANALYSIS_MODULES
from engine.data import DATA_MODULES, mod_pricing
from engine.xlsx import write_plain_table

# User-owned tabs. Python renders them once and never writes them again — refresh
# leaves them untouched so user edits survive.
INPUTS_TAB = "Inputs"
NOTES_TAB = "Notes"

# Reference tabs.
REF_PROVENANCE_TAB = "Ref_Provenance"
REF_LEGEND_TAB = "Ref_Legend"

# Single-cell runtime inputs, laid out as (label, named-range name) down column A/B
# of the Inputs tab. Each value cell is exposed as a workbook-scope named range so
# Analysis formulas reference e.g. ``Input_Horizon`` instead of ``Inputs!$B$2``.
INPUT_ROWS: list[tuple[str, str | None]] = [
    ("Horizon (years)", "Input_Horizon"),
    ("HYSA rate (%)", "Input_HYSA_Rate"),
    ("Down payment ($)", "Input_Down_Payment"),
    ("Trade-in value ($)", "Input_Trade_In"),
    ("State", "Input_State"),
    ("ZIP", "Input_ZIP"),
    ("Home charging (Y/N)", "Input_Home_Charging"),
    ("Labor rate ($/hr)", "Input_Labor_Rate"),
]

# Default values written into the value cell (column B) at render for the inputs
# that need a live starting number. Input_Labor_Rate seeds the aftermarket-labor
# math on Analysis_TrimPath (labor = install hours x this rate); the other inputs
# stay blank for the user to fill.
INPUT_DEFAULTS: dict[str, object] = {
    "Input_Labor_Rate": 120,
}

# The Target_Build user-input table on Inputs (this design, IC #19). Two columns:
# the taxonomy display name (joins Features[Feature] / ModPricing[Feature]) and a
# Yes/No Include flag the user marks. Seeded from mod-pricing coverage.
TARGET_BUILD_TABLE = "Target_Build"
TARGET_BUILD_COLUMNS = ["Feature", "Include"]


def _render_inputs(ws: Worksheet) -> None:
    ws["A1"] = "Input"
    ws["B1"] = "Value"
    for offset, (label, name) in enumerate(INPUT_ROWS, start=2):
        ws.cell(row=offset, column=1, value=label)
        if name in INPUT_DEFAULTS:
            ws.cell(row=offset, column=2, value=INPUT_DEFAULTS[name])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    _render_target_build(ws, start_row=len(INPUT_ROWS) + 3)


def _render_target_build(ws: Worksheet, start_row: int) -> None:
    """Render the Target_Build section below the single-cell inputs block.

    A bold section header + a wrapped explanatory subheader, then a provenance-free
    Excel Table seeded with every mod-pricing-covered feature (Include="No"), with a
    Yes/No data-validation dropdown on the Include column. Seeding from mod-pricing
    coverage (not the full taxonomy) is deliberate: a feature with no aftermarket row
    has no aftermarket landed-cost path, so it should not be offered as a target.
    """
    header_row = start_row
    ws.cell(row=header_row, column=1, value="Target build").font = Font(bold=True)

    sub = ws.cell(
        row=header_row + 1,
        column=1,
        value=(
            'Mark Include = "Yes" for each capability your build must reach. Every '
            "trim path's landed cost is computed to hit exactly these. Aftermarket "
            "pricing is only defined for the features listed here (the ones with a "
            "practical aftermarket); factory-only trim hardware still shows as "
            "standard-embedded where a trim includes it."
        ),
    )
    sub.alignment = Alignment(wrap_text=True, vertical="top")

    # Seed one row per mod-pricing-covered feature. mod_pricing.load() emits the
    # display name in its Feature column (COLUMNS[0]) — the same id->name map the two
    # data tabs join by — so reuse it rather than re-deriving the id->name mapping.
    feature_names = [row[0] for row in mod_pricing.load()]
    rows = [[name, "No"] for name in feature_names]

    table_start = header_row + 3
    write_plain_table(
        ws, TARGET_BUILD_TABLE, TARGET_BUILD_COLUMNS, rows, start_row=table_start
    )

    # Yes/No dropdown on the Include column (column B) across the data rows, so the
    # cell reads like a checkbox and can't hold a typo the formulas would miss.
    first_data = table_start + 1
    last_data = table_start + len(rows)
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    dv.add(f"B{first_data}:B{last_data}")
    ws.add_data_validation(dv)


def _render_notes(ws: Worksheet) -> None:
    ws["A1"] = "Notes"
    ws["A2"] = "Free-form space for your own notes. Python never writes this tab."
    ws.column_dimensions["A"].width = 60


def _render_ref_provenance(ws: Worksheet) -> None:
    for col, name in enumerate(["Tab", "Column", "Value", "Source", "As_Of_Date"], start=1):
        ws.cell(row=1, column=col, value=name)
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["C"].width = 24


def _render_ref_legend(ws: Worksheet) -> None:
    lines = [
        ("How to read this workbook", ""),
        ("", ""),
        ("Ownership model", "Each tab is owned by exactly one party."),
        ("Inputs, Notes", "You own these. Edit freely. `engine refresh` never touches them."),
        ("Data_*", "Python owns these. `engine refresh` rewrites their rows from source data."),
        ("Analysis_*", "Excel formulas compute these from Data_* tables + your Inputs. No Python."),
        ("Ref_*", "Reference: Ref_Provenance lists every value's source; Ref_Legend is this tab."),
        ("", ""),
        ("Discipline", ""),
        ("Excel Tables", "Data lives in named Tables (e.g. Trims[MSRP]) so refresh never breaks formulas."),  # noqa: E501
        ("Named ranges", "Single-cell Inputs are named ranges (e.g. Input_Horizon) formulas reference."),  # noqa: E501
        ("Provenance", "Every Data_* row carries Source + As_Of_Date; Ref_Provenance consolidates them."),  # noqa: E501
    ]
    for offset, (left, right) in enumerate(lines, start=1):
        ws.cell(row=offset, column=1, value=left)
        ws.cell(row=offset, column=2, value=right)
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 90


def _add_named_ranges(wb: Workbook) -> None:
    # Single-cell input named ranges (the runtime-parameter pattern).
    for offset, (_label, name) in enumerate(INPUT_ROWS, start=2):
        if name is None:
            continue
        ref = f"{INPUTS_TAB}!$B${offset}"
        wb.defined_names[name] = DefinedName(name=name, attr_text=ref)

    # At least one named range referencing a Table, demonstrating the pattern that
    # Analysis formulas will use to pull structured columns out of Data_* tables.
    wb.defined_names["Trims_MSRP"] = DefinedName(name="Trims_MSRP", attr_text="Trims[MSRP]")


def build_workbook() -> Workbook:
    """Build and return the full report workbook (in memory; caller saves it)."""
    wb = Workbook()
    # Drop the default sheet; we create every tab explicitly and in order.
    wb.remove(wb.active)

    _render_inputs(wb.create_sheet(INPUTS_TAB))
    _render_notes(wb.create_sheet(NOTES_TAB))

    for module in DATA_MODULES:
        module.render(wb.create_sheet(module.TAB_NAME))

    for module in ANALYSIS_MODULES:
        module.setup_formulas(wb.create_sheet(module.TAB_NAME))

    _render_ref_provenance(wb.create_sheet(REF_PROVENANCE_TAB))
    _render_ref_legend(wb.create_sheet(REF_LEGEND_TAB))

    _add_named_ranges(wb)
    return wb
