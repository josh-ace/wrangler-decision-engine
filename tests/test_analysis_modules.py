"""Unit tests for the Analysis_* module plug-in pattern.

One test case per analysis module (parametrized). The contract: declare the tab
and expose ``setup_formulas(ws)`` where a future IC writes Excel formulas.
"""

from __future__ import annotations

import re

import pytest
from openpyxl import Workbook

from engine.analysis import ANALYSIS_MODULES, trim_path

MODULE_IDS = [m.TAB_NAME for m in ANALYSIS_MODULES]


@pytest.mark.parametrize("module", ANALYSIS_MODULES, ids=MODULE_IDS)
def test_analysis_module_declares_tab(module):
    assert module.TAB_NAME.startswith("Analysis_")
    assert module.TITLE


@pytest.mark.parametrize("module", ANALYSIS_MODULES, ids=MODULE_IDS)
def test_analysis_module_setup_formulas_runs(module):
    wb = Workbook()
    ws = wb.active
    module.setup_formulas(ws)  # every module writes at least its title
    assert ws["A1"].value == module.TITLE


# --- Analysis_TrimPath formula-string assertions (IC #19) -------------------
# openpyxl doesn't calculate, so these assert the formula *text* the module writes:
# the structured-reference discipline (no cell addresses) and the specific formulas
# the design fixes. Complements the arithmetic re-derivation in test_e2e.py.


def _trim_path_formulas():
    wb = Workbook()
    ws = wb.active
    trim_path.setup_formulas(ws)
    formulas = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str) and c.value.startswith("=")
    ]
    return ws, formulas


def test_trim_path_formulas_use_structured_refs_only():
    _ws, formulas = _trim_path_formulas()
    assert formulas, "expected formulas to be written"
    for f in formulas:
        assert not re.search(r"\$[A-Z]+\$[0-9]", f), f


def test_trim_path_writes_both_tables_with_row_counts():
    ws, _formulas = _trim_path_formulas()
    assert trim_path.SUMMARY_TABLE in ws.tables
    assert trim_path.DETAIL_TABLE in ws.tables

    def rows(table):
        m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", ws.tables[table].ref)
        return int(m.group(2)) - int(m.group(1))  # data rows (header excluded)

    assert rows(trim_path.SUMMARY_TABLE) == 28
    assert rows(trim_path.DETAIL_TABLE) == 560


def test_trim_path_key_formulas_match_design():
    _ws, formulas = _trim_path_formulas()
    joined = "\n".join(formulas)

    # Base_MSRP is a keyed SUMIFS into Trims (not positional, survives refresh).
    assert (
        "=SUMIFS(Trims[MSRP], Trims[Trim], [@Trim], "
        "Trims[Body], [@Body], Trims[Powertrain], [@Powertrain])" in formulas
    )
    # Landed_Cost sums MSRP + aftermarket parts + aftermarket labor.
    assert "=[@Base_MSRP] + [@Aftermarket_Cost] + [@Aftermarket_Labor_Cost]" in formulas
    # Delta is a structured self-reference over the whole Landed_Cost column.
    assert "=[@Landed_Cost] - MIN(TrimPathSummary[Landed_Cost])" in formulas
    # Aftermarket labor multiplies install hours by the Input_Labor_Rate named range.
    assert "Input_Labor_Rate" in joined
    # Standard_On_Trim is delimiter-bracketed (guards Sport-in-Sport S), not a naive
    # SEARCH of the bare trim name.
    assert '", "&[@Trim]&","' in joined
    assert "ISNUMBER(SEARCH(" in joined


def test_trim_path_factory_option_is_na_literal():
    ws, _formulas = _trim_path_formulas()
    # The Factory_Option_Price / Factory_Options_Cost cells are the explicit MVP
    # placeholder, not a fabricated number.
    values = {c.value for row in ws.iter_rows() for c in row}
    assert trim_path.FACTORY_OPTION_NA in values
