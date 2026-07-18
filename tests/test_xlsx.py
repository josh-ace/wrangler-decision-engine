"""Unit tests for the shared Excel-Table / named-range helpers."""

from __future__ import annotations

from openpyxl import Workbook

from engine.xlsx import PROVENANCE_COLUMNS, clear_data_table, write_data_table, write_plain_table


def test_write_data_table_appends_provenance_columns():
    wb = Workbook()
    ws = wb.active
    write_data_table(ws, "Demo", ["A", "B"], rows=[])
    header = [c.value for c in ws[1]]
    assert header == ["A", "B", *PROVENANCE_COLUMNS]
    assert "Demo" in ws.tables


def test_write_data_table_emits_placeholder_row_when_empty():
    wb = Workbook()
    ws = wb.active
    write_data_table(ws, "Demo", ["A", "B"], rows=[])
    # Header + exactly one (blank) placeholder row so the table ref is valid.
    assert ws.tables["Demo"].ref == "A1:D2"


def test_write_data_table_writes_real_rows():
    wb = Workbook()
    ws = wb.active
    write_data_table(ws, "Demo", ["A", "B"], rows=[[1, 2, "src", "2026-01-01"]])
    assert [c.value for c in ws[2]] == [1, 2, "src", "2026-01-01"]


def test_clear_data_table_removes_table_and_cells():
    wb = Workbook()
    ws = wb.active
    write_data_table(ws, "Demo", ["A", "B"], rows=[])
    clear_data_table(ws, "Demo")
    assert "Demo" not in ws.tables
    assert ws["A1"].value is None


def test_formula_columns_get_calculated_column_formula():
    """Regression test: formula cells inside a table only resolve [@X] references
    in Excel when the TableColumn carries calculatedColumnFormula. Without it,
    Excel treats each cell as a standalone formula and [@Trim] fails. The fix
    auto-detects formula columns (first data row starts with '=') and attaches
    the marker; this asserts it fires end-to-end.
    """
    wb = Workbook()
    ws = wb.active
    formula = '=[@A]&"-"&[@B]'
    write_plain_table(ws, "Demo", ["A", "B", "C"], rows=[["x", "y", formula]])
    table = ws.tables["Demo"]
    a_col, b_col, c_col = table.tableColumns
    # Literal columns get no calculated formula.
    assert a_col.calculatedColumnFormula is None
    assert b_col.calculatedColumnFormula is None
    # Formula column has one, matching the cell body (without leading '=').
    assert c_col.calculatedColumnFormula is not None
    assert c_col.calculatedColumnFormula.attr_text == '[@A]&"-"&[@B]'
