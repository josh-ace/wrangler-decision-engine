"""End-to-end tests exercising the whole pipeline through the CLI.

These are the tests future ICs extend: when a module fills in real data or
formulas, its assertions grow here alongside the smoke test.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from click.testing import CliRunner
from openpyxl import load_workbook

from engine.analysis import ANALYSIS_MODULES, trim_path
from engine.cli import main
from engine.data import DATA_MODULES, features, mod_pricing, trims
from engine.workbook import (
    INPUTS_TAB,
    NOTES_TAB,
    REF_LEGEND_TAB,
    REF_PROVENANCE_TAB,
    TARGET_BUILD_TABLE,
)

_FEATURES_JSON = Path(__file__).resolve().parents[1] / "data" / "features.json"

# Authoritative expected tab set, derived from the same registries the renderer
# uses — add a module and this expectation updates itself.
EXPECTED_TABS = [
    INPUTS_TAB,
    NOTES_TAB,
    *[m.TAB_NAME for m in DATA_MODULES],
    *[m.TAB_NAME for m in ANALYSIS_MODULES],
    REF_PROVENANCE_TAB,
    REF_LEGEND_TAB,
]


def test_render_produces_all_tabs_tables_and_named_range(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"

    result = runner.invoke(main, ["render", "-o", str(out)])
    assert result.exit_code == 0, result.output
    assert out.exists()

    wb = load_workbook(out)

    # Every expected tab exists.
    assert wb.sheetnames == EXPECTED_TABS

    # Every Data_* tab has its Excel Table set up.
    for module in DATA_MODULES:
        ws = wb[module.TAB_NAME]
        assert module.TABLE_NAME in ws.tables, f"{module.TAB_NAME} missing its table"

    # At least one named range exists (and the table-referencing one is present).
    assert wb.defined_names
    assert "Trims_MSRP" in wb.defined_names


def _read_table_rows(ws, header):
    """Return the Data_* tab's rows (below the header) as a list of dict rows keyed
    by column name, reading straight from the worksheet cells."""
    names = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(names, r)))
    assert names[: len(header)] == header
    return rows


def test_data_trims_has_real_rows_with_values_and_provenance(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"
    assert runner.invoke(main, ["render", "-o", str(out)]).exit_code == 0

    wb = load_workbook(out)
    ws = wb[trims.TAB_NAME]
    rows = _read_table_rows(ws, trims.COLUMNS)

    # One row per (trim x body x configuration) — 28 across both body styles.
    assert len(rows) == 28

    # Every row carries populated provenance.
    for row in rows:
        assert row["Source"] == "order_guide_2026"
        assert row["As_Of_Date"] == "2025-08-06"

    # Spot-check a known 2-door value: Sport V6 6MT MSRP = $33,785.
    sport_2dr = next(
        r for r in rows
        if r["Trim"] == "Sport" and r["Body"] == "JL 2-door" and r["Powertrain"] == "V6 6MT"
    )
    assert sport_2dr["MSRP"] == 33785
    assert sport_2dr["Invoice"] is None

    # Spot-check a known 4-door value: Rubicon V6 6MT MSRP = $49,270.
    rubicon_4dr = next(
        r for r in rows
        if r["Trim"] == "Rubicon" and r["Body"] == "JLU 4-door" and r["Powertrain"] == "V6 6MT"
    )
    assert rubicon_4dr["MSRP"] == 49270


def test_data_features_has_real_taxonomy_rows(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"
    assert runner.invoke(main, ["render", "-o", str(out)]).exit_code == 0

    taxonomy = json.loads(_FEATURES_JSON.read_text(encoding="utf-8"))

    wb = load_workbook(out)
    ws = wb[features.TAB_NAME]
    rows = _read_table_rows(ws, features.COLUMNS)

    # Row count matches the taxonomy exactly.
    assert len(rows) == len(taxonomy["features"])

    # Feature ids (slugs) are all unique — no duplicate features in the taxonomy.
    ids = [f["id"] for f in taxonomy["features"]]
    assert len(ids) == len(set(ids))

    # Every rendered row carries curator-attributed provenance and null prices.
    for row in rows:
        assert "curator" in row["Source"]
        assert row["As_Of_Date"] == "2026-07-18"
        assert row["Factory_Option_Price"] is None
        assert row["Aftermarket_Price"] is None

    by_feature = {row["Feature"]: row for row in rows}

    # Spot-check known off-road features and their standard-on availability.
    assert "Rubicon" in by_feature["Rear Locker"]["Standard_On"]
    rock_rails = by_feature["Rock Rails"]["Standard_On"]
    assert "Willys" in rock_rails and "Rubicon" in rock_rails


def test_data_mod_pricing_has_real_rows_joined_to_features(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"
    assert runner.invoke(main, ["render", "-o", str(out)]).exit_code == 0

    wb = load_workbook(out)

    # Every ModPricing Feature must be a real Data_Features name (the join key).
    feature_rows = _read_table_rows(wb[features.TAB_NAME], features.COLUMNS)
    feature_names = {row["Feature"] for row in feature_rows}

    ws = wb[mod_pricing.TAB_NAME]
    rows = _read_table_rows(ws, mod_pricing.COLUMNS)

    # Real rows exist (not the empty-placeholder scaffold).
    assert len(rows) > 0

    for row in rows:
        # Feature name resolves to a feature in Data_Features.
        assert row["Feature"] in feature_names, row["Feature"]
        # Parts cost + install hours are populated with sane values.
        assert row["Parts_Cost"] is not None and row["Parts_Cost"] > 0
        assert row["Install_Hours"] is not None and row["Install_Hours"] >= 0
        # Provenance is populated per-row.
        assert "curator" in row["Source"]
        assert row["As_Of_Date"] == "2026-07-18"

    # Spot-check a headline mod-target join: the rear locker is priced.
    by_feature = {row["Feature"]: row for row in rows}
    assert by_feature["Rear Locker"]["Parts_Cost"] > 0


def test_refresh_preserves_user_edits(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"

    assert runner.invoke(main, ["render", "-o", str(out)]).exit_code == 0

    # User edits a cell on the Inputs tab (the HYSA-rate value cell).
    wb = load_workbook(out)
    wb[INPUTS_TAB]["B3"] = 4.75
    wb.save(out)

    # Refresh rewrites Data_* tabs only.
    result = runner.invoke(main, ["refresh", str(out)])
    assert result.exit_code == 0, result.output

    # The user's edit survives, and the data tables are still intact.
    wb2 = load_workbook(out)
    assert wb2[INPUTS_TAB]["B3"].value == 4.75
    for module in DATA_MODULES:
        assert module.TABLE_NAME in wb2[module.TAB_NAME].tables


def test_refresh_missing_file_errors_cleanly(tmp_path):
    runner = CliRunner()
    result = runner.invoke(main, ["refresh", str(tmp_path / "nope.xlsx")])
    assert result.exit_code != 0
    assert "No workbook to refresh" in result.output


# ---------------------------------------------------------------------------
# Analysis_TrimPath (IC #19). openpyxl does not calculate formulas, so — per the
# design doc's recommendation — the landed-cost assertions re-derive the five-feature
# model in Python from the same source tables the workbook renders and assert the
# arithmetic against the worked example, while a separate test asserts the formula
# strings use structured references only.
# ---------------------------------------------------------------------------

# The worked example's target build: five capabilities marked Include="Yes"
# (taxonomy display names, matching Target_Build / ModPricing / Features join keys).
TARGET_BUILD = [
    "Rear Locker",
    "35-Inch Tires (Factory)",
    "Rock Rails",
    "Steel Front & Rear Bumpers",
    "Factory Warn Winch",
]
LABOR_RATE = 120  # Input_Labor_Rate default

# Worked-example landed costs (JLU 4-door, V6 8AT) from the design doc.
WORKED_EXAMPLE = {
    ("Sport", "JLU 4-door", "V6 8AT"): 48_890,
    ("Willys", "JLU 4-door", "V6 8AT"): 54_390,
    ("Rubicon", "JLU 4-door", "V6 8AT"): 58_350,
}


def _table_ref_bounds(table):
    """Return (first_data_row, last_data_row) for an openpyxl Table (header excluded)."""
    m = re.match(r"[A-Z]+(\d+):[A-Z]+(\d+)", table.ref)
    header_row, last_row = int(m.group(1)), int(m.group(2))
    return header_row + 1, last_row


def _standard_on_match(trim: str, standard_on: str) -> bool:
    """Replicate the sheet's delimiter-bracketed, case-insensitive SEARCH so 'Sport'
    does not false-match inside 'Sport S'. This is the exact membership test the
    Standard_On_Trim formula encodes."""
    haystack = f", {standard_on or ''},".lower()
    needle = f", {trim},".lower()
    return needle in haystack


def _landed_from_workbook(wb, config, target, rate):
    """Re-derive a config's landed cost from the rendered Data_* tables: MSRP plus,
    for each included feature not standard-embedded on the trim, parts + hours*rate."""
    trim, body, powertrain = config

    trim_rows = _read_table_rows(wb[trims.TAB_NAME], trims.COLUMNS)
    msrp = next(
        r["MSRP"]
        for r in trim_rows
        if (r["Trim"], r["Body"], r["Powertrain"]) == config
    )

    feat_rows = _read_table_rows(wb[features.TAB_NAME], features.COLUMNS)
    standard_on = {r["Feature"]: r["Standard_On"] for r in feat_rows}

    mod_rows = _read_table_rows(wb[mod_pricing.TAB_NAME], mod_pricing.COLUMNS)
    mods = {r["Feature"]: (r["Parts_Cost"], r["Install_Hours"]) for r in mod_rows}

    landed = msrp
    for feature in target:
        if _standard_on_match(trim, standard_on[feature]):
            continue  # standard-embedded — cost already in MSRP
        parts, hours = mods[feature]
        landed += parts + hours * rate
    return landed


def _render(tmp_path):
    runner = CliRunner()
    out = tmp_path / "wrangler.xlsx"
    assert runner.invoke(main, ["render", "-o", str(out)]).exit_code == 0
    return out


def test_trim_path_landed_costs_match_worked_example(tmp_path):
    """Assertions 1–3: Sport/Willys/Rubicon landed within $50 of the worked example."""
    wb = load_workbook(_render(tmp_path))
    for config, expected in WORKED_EXAMPLE.items():
        landed = _landed_from_workbook(wb, config, TARGET_BUILD, LABOR_RATE)
        assert abs(landed - expected) <= 50, f"{config}: {landed} vs {expected}"


def test_trim_path_calibration_guard_rail(tmp_path):
    """Assertion 4: Sport-mods landed is at least $5,000 under at-target Rubicon."""
    wb = load_workbook(_render(tmp_path))
    sport = _landed_from_workbook(
        wb, ("Sport", "JLU 4-door", "V6 8AT"), TARGET_BUILD, LABOR_RATE
    )
    rubicon = _landed_from_workbook(
        wb, ("Rubicon", "JLU 4-door", "V6 8AT"), TARGET_BUILD, LABOR_RATE
    )
    assert rubicon - sport >= 5_000, f"gap only {rubicon - sport}"


def test_trim_path_no_hardcoded_cell_addresses(tmp_path):
    """Assertion 5: every formula on Analysis_TrimPath uses structured refs / named
    ranges — zero `$COL$ROW` hits."""
    wb = load_workbook(_render(tmp_path))
    ws = wb[trim_path.TAB_NAME]
    hits = [
        c.value
        for row in ws.iter_rows()
        for c in row
        if isinstance(c.value, str)
        and c.value.startswith("=")
        and re.search(r"\$[A-Z]+\$[0-9]", c.value)
    ]
    assert hits == []


def test_trims_msrp_named_range_still_resolves(tmp_path):
    """Assertion 6: adding this tab leaves the pre-existing Trims_MSRP named range
    pointing at Trims[MSRP]."""
    wb = load_workbook(_render(tmp_path))
    assert "Trims_MSRP" in wb.defined_names
    assert wb.defined_names["Trims_MSRP"].attr_text == "Trims[MSRP]"


def test_standard_on_delimiter_safety():
    """Assertion 7: 'Sport' is not counted standard for a feature standard on
    'Sport S, Willys, Rubicon, Rubicon X' (the Sport-in-Sport S false positive the
    bracketed SEARCH prevents), while 'Sport S' itself matches."""
    standard_on = "Sport S, Willys, Rubicon, Rubicon X"
    assert _standard_on_match("Sport", standard_on) is False
    assert _standard_on_match("Sport S", standard_on) is True
    assert _standard_on_match("Rubicon", standard_on) is True
    assert _standard_on_match("Rubicon X", standard_on) is True


def test_trim_path_has_two_tables_with_expected_row_counts(tmp_path):
    """Assertion 8: TrimPathSummary (28 rows) and TrimPathDetail (560 rows) exist."""
    wb = load_workbook(_render(tmp_path))
    ws = wb[trim_path.TAB_NAME]
    assert trim_path.SUMMARY_TABLE in ws.tables
    assert trim_path.DETAIL_TABLE in ws.tables

    first, last = _table_ref_bounds(ws.tables[trim_path.SUMMARY_TABLE])
    assert last - first + 1 == 28

    first, last = _table_ref_bounds(ws.tables[trim_path.DETAIL_TABLE])
    assert last - first + 1 == 28 * 20  # 560


def test_refresh_preserves_target_build_selections(tmp_path):
    """Assertion 9: marking Target_Build Include values and refreshing leaves them
    untouched (Inputs is never rewritten by refresh)."""
    runner = CliRunner()
    out = _render(tmp_path)

    wb = load_workbook(out)
    ws = wb[INPUTS_TAB]
    first, last = _table_ref_bounds(ws.tables[TARGET_BUILD_TABLE])
    # Mark the first two features Include="Yes" (Include is column B of the table).
    ws.cell(row=first, column=2, value="Yes")
    ws.cell(row=first + 1, column=2, value="Yes")
    marked_features = [ws.cell(row=first, column=1).value, ws.cell(row=first + 1, column=1).value]
    wb.save(out)

    assert runner.invoke(main, ["refresh", str(out)]).exit_code == 0

    wb2 = load_workbook(out)
    ws2 = wb2[INPUTS_TAB]
    assert ws2.cell(row=first, column=2).value == "Yes"
    assert ws2.cell(row=first + 1, column=2).value == "Yes"
    # And the features they were attached to are unchanged (row identity preserved).
    assert ws2.cell(row=first, column=1).value == marked_features[0]
    assert ws2.cell(row=first + 1, column=1).value == marked_features[1]
